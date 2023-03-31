from openpyxl.drawing.image import Image
from os import getcwd, path, makedirs
from PIL import Image
from win32com import client
import openpyxl
import os
import qrcode
import subprocess


class GerarEtiqueta:
    def __init__(self, material, nf, bancada, lote, deposito, caixa, qtd, seriais):
        self.path = getcwd()
        self.path = str(rf'{self.path}\assets')
        self.material = material
        self.nf = nf
        self.bancada = bancada
        self.lote = lote
        self.deposito = deposito
        self.caixa = caixa
        self.qtd = qtd
        self.seriais = seriais

        # Definindo os diretórios dos arquivos da etiqueta
        self.qr_file = rf'{self.path}\temp\qrcode.png'
        self.etq_modelo = rf'{self.path}\etiqueta.xlsx'
        self.etq_final = rf'{self.path}\temp\etq.xlsx'
        self.etq_pdf = rf'{self.path}\temp\etq.pdf'
        self.pdf_printer = rf'{self.path}\PDFtoPrinter.exe'

        # Criar pasta temporária para salvar os arquivos
        if not path.isdir(rf'{self.path}\temp'):
            makedirs(rf'{self.path}\temp')

        self.gerar_qrCode()
        self.inserir_dados()
        self.convert_pdf()
        self.impressao()
        self.apagar_temp()

    # Gerar o QRCODE com os seriais da leitura
    def gerar_qrCode(self):
        qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=10, border=0)

        qr.add_data(self.seriais)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        self.qr_file = str(self.qr_file)
        img.save(self.qr_file)

        # Redimensionar a imagem do QRCODE
        img = Image.open(self.qr_file)
        img = img.resize((210, 210), Image.LANCZOS)
        img.save(self.qr_file)

    # Inserir o QRCODE na planilha da etiqueta
    def inserir_dados(self):

        # Abrir a planilha e inserir o cabeçalho da etiqueta
        wb = openpyxl.load_workbook(filename=self.etq_modelo)
        ws = wb.worksheets[0]
        img = openpyxl.drawing.image.Image(self.qr_file)
        ws['E3'] = self.material
        ws['E5'] = self.nf
        ws['G5'] = self.lote
        ws['E8'] = self.deposito
        ws['G8'] = self.caixa
        ws['H8'] = self.qtd

        # Inserir o QRCODE na planilha
        img.anchor = 'C10'
        ws.add_image(img)
        wb.save(self.etq_final)

    # Converter a etiqueta em PDF
    def convert_pdf(self):
        excel = client.Dispatch('Excel.Application')
        excel.DisplayAlerts = False
        sheets = excel.Workbooks.Open(self.etq_final)
        ws = sheets.Worksheets[0]
        ws.ExportAsFixedFormat(0, self.etq_pdf)
        excel.Quit()

    # Imprimir o PDF da etiqueta gerada
    def impressao(self):
        comando = "{} {} {}".format(self.pdf_printer, self.etq_pdf, '/s')
        subprocess.call(comando, shell=True)

    # Apagar a pasta temp
    def apagar_temp(self):
        temp_folder = rf'{self.path}\temp'

        for file in os.listdir(temp_folder):
            file_path = os.path.join(temp_folder, file)
            if os.path.isfile(file_path):
                os.unlink(file_path)

        os.rmdir(temp_folder)
